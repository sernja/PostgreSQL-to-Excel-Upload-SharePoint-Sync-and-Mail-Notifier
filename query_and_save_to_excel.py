import psycopg2
from openpyxl import Workbook

# Connect to the PostgreSQL database
def query(host, dbname, user, password, port, date_from, date_to, output_excel_file):
    try:
        connection = psycopg2.connect(
            host=host,
            database=dbname,
            user=user,
            password=password,
            port=port
        )
        cursor = connection.cursor()

        # Example: Fetch all rows from a table
        cursor.execute(f'''
                    SELECT * from table
    WHERE date >= '{date_from}'::date AND date <= '{date_to}'::date
    ORDER BY date;
                    ''')
        rows = cursor.fetchall()

        # Get column names
        colnames = [desc[0] for desc in cursor.description]

        # Create a workbook and add a worksheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Query Results"

        # Write column headers
        for col_num, col_name in enumerate(colnames, start=1):
            sheet.cell(row=1, column=col_num, value=col_name)

        # Write data rows
        for row_num, row_data in enumerate(rows, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        # Save to Excel file
        workbook.save(output_excel_file)
        print(f"Query results have been saved to {output_excel_file}")

    except Exception as error:
        print("Error while connecting to PostgreSQL:", error)
    finally:
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection is closed")
